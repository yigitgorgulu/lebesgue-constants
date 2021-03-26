import argparse

from wolframclient.evaluation import WolframLanguageSession
from wolframclient.language import wl, wlexpr

from openpyxl import load_workbook, Workbook

import matplotlib.pyplot as plt
import numpy as np

# kernel path for wolfram client
KERNEL_PATH = 'D:\\Wolfram Research\\Wolfram Engine\\12.1\WolframKernel.exe'

# parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument('a', help='the alpha value for the set')
parser.add_argument('l', help='initial l_1 (reciprocal is taken)')
parser.add_argument('s', help='the step')
parser.add_argument('-m', action='store_true', help='whether to include m')
parser.add_argument('-d', action='store_true', help='whether to display or save file')
args = parser.parse_args()

file = '../output/values.xlsx'
wb = load_workbook(file)

ws = wb.create_sheet('alpha_a{}s{}l{}{}'.format(args.a, args.s, args.l, 'm' if args.m else ''))

# access the specified kernel
with WolframLanguageSession(KERNEL_PATH) as session:
	if args.m:
		session.evaluate(wlexpr('alpha = {};'.format(args.a)))
		session.evaluate(wlexpr('l[1] = 1 / {};'.format(args.l)))
		session.evaluate(wlexpr('l[n_] := l[n - 1] ^ alpha;'))
		
		session.evaluate(wlexpr('k[0] = {0, 1};'))
		session.evaluate(wlexpr('k[n_] := Union[Table[N[k[n - 1][[i]]], {i, 1, Length[k[n - 1]]}], Table[N[k[n - 1][[i]]], {i, 1, Length[k[n - 1]], 2}] + l[n], Table[N[k[n - 1][[i]]], {i, 2, Length[k[n - 1]], 2}] - l[n]];'))
		
		session.evaluate(wlexpr('s = {};'.format(args.s)))
		session.evaluate(wlexpr('set = k[s - 1]'))
		set = session.evaluate(wlexpr('N[set]'))
		session.evaluate(wlexpr('next = k[s];'))
		next = session.evaluate(wlexpr('N[next]'))
		lebesgue = session.evaluate(wlexpr('Table[N[Sum[Product[If[k == j, 1, Abs[(next[[i]] - set[[k]]) / (set[[j]] - set[[k]])]], {k, 1, 2 ^ s}], {j, 1, 2 ^ s}]], {i, 1, 2 ^ (s + 1)}]'))

	else:
		session.evaluate(wlexpr('alpha = {};'.format(args.a)))
		session.evaluate(wlexpr('l[1] = 1 / {};'.format(args.l)))
		session.evaluate(wlexpr('l[n_] := l[n - 1] ^ alpha;'))
		
		session.evaluate(wlexpr('k[0] = {0, 1};'))
		session.evaluate(wlexpr('k[n_] := Union[Table[N[k[n - 1][[i]]], {i, 1, Length[k[n - 1]]}], Table[N[k[n - 1][[i]]], {i, 1, Length[k[n - 1]], 2}] + l[n], Table[N[k[n - 1][[i]]], {i, 2, Length[k[n - 1]], 2}] - l[n]];'))
		
		session.evaluate(wlexpr('findxm[n_]:=N[1 + Sum[l[i] * (-1) ^ i, {i, 1, n}]];'))
		
		session.evaluate(wlexpr('s = {};'.format(args.s)))
		session.evaluate(wlexpr('set = k[s - 1]'))
		set = session.evaluate(wlexpr('N[set]'))
		session.evaluate(wlexpr('next = k[s];'))
		next = session.evaluate(wlexpr('N[next]'))
		session.evaluate(wlexpr('xm = findxm[s - 1];'))
		session.evaluate(wlexpr('m = Position[set, xm][[1, 1]];'))
		lebesgue = session.evaluate(wlexpr('Table[N[Sum[If[j == m, 0, Product[If[k == m || k == j, 1, Abs[(next[[i]] - set[[k]]) / (set[[j]] - set[[k]])]], {k, 1, 2 ^ s}]], {j, 1, 2 ^ s}]], {i, 1, 2 ^ (s + 1)}]'))
		
	# calculate amount of rows
	rows = 2 ** (int(args.s) + 1)
	
	# create dictionary from keys and values
	results = dict(zip(next, lebesgue))
	
	# write the values
	for i in range(rows):
		ws['A{}'.format(i + 1)].value = next[i]
		ws['B{}'.format(i + 1)].value = lebesgue[i]
	
	# save the file
	wb.save(file)	
	
	# generate and show plot
	plt.plot(next, lebesgue)
	plt.title('a = {}, s = {}, l_1 = {}{}'.format(args.a, args.s, args.l, ', xm included' if args.m else ''))
	if args.d:
		plt.show()
	else:
		plt.savefig('../output/alpha/alpha_a{}s{}l{}{}.png'.format(args.a, args.s, args.l, 'm' if args.m else ''))
		
	'''
	results = dict(zip(next, lebesgue))
	for i, j in results.items():
		print('{}: {}'.format(i, j))
	'''