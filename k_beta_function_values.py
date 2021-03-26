import argparse

from wolframclient.evaluation import WolframLanguageSession
from wolframclient.language import wl, wlexpr

from openpyxl import load_workbook, Workbook

import matplotlib.pyplot as plt
import numpy as np

# kernel path for wolfram client
KERNEL_PATH = 'D:\\Wolfram Research\\Wolfram Engine\\12.1\WolframKernel.exe'

# parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument('b', help='the beta value for the set', type=int)
parser.add_argument('s', help='the step', type=int)
args = parser.parse_args()

# check whether an xlsx file was provided and open it (or a new one)
file = '../output/values.xlsx'
wb = load_workbook(file)

# access correct sheet
ws = wb['beta={}'.format(args.b)]

# access the specified kernel
with WolframLanguageSession(KERNEL_PATH) as session:
	session.evaluate(wlexpr('b = {};'.format(args.b)))

	session.evaluate(wlexpr('cantor[0] = {0, 1};'))
	session.evaluate(wlexpr('cantor[n_] := Join[cantor[n - 1] / b, (b - 1 + cantor[n - 1]) / b];'))

	session.evaluate(wlexpr('s = {};'.format(args.s)))
	session.evaluate(wlexpr('set = cantor[s - 1]'))
	set = session.evaluate(wlexpr('N[set]'))
	session.evaluate(wlexpr('next = cantor[s];'))
	next = session.evaluate(wlexpr('N[next]'))

	session.evaluate('l[k_, x_] := Product[If[((j == k) || (set[[j]] == x)), 1, Abs[(x - set[[j]]) / (set[[k]] - set[[j]])]], {j, 1, 2 ^ s}];')

	max_values = session.evaluate(wlexpr('Table[Max[Table[N[l[k, next[[i]]]], {i, 1, 2 ^ (s + 1)}]], {k, 1, 2 ^ s}]'))

	for i in range(2 ** args.s):
		if i < 25: # single letter columns
			ws['{}{}'.format(chr(ord('B') + i), (args.s + 1))] = max_values[i]
		elif i < 51: # columns starting with A
			ws['A{}{}'.format(chr(ord('A') + i - 25), (args.s + 1))] = max_values[i]
		elif i < 77: # columns starting with B
			ws['B{}{}'.format(chr(ord('A') + i - 51), (args.s + 1))] = max_values[i]
		elif i < 103: # columns starting with C
			ws['C{}{}'.format(chr(ord('A') + i - 77), (args.s + 1))] = max_values[i]
		elif i < 129: # columns starting with D
			ws['D{}{}'.format(chr(ord('A') + i - 103), (args.s + 1))] = max_values[i]

	wb.save(file)