import argparse

from wolframclient.evaluation import WolframLanguageSession
from wolframclient.language import wl, wlexpr

from openpyxl import load_workbook, Workbook

import matplotlib.pyplot as plt
import numpy as np

# kernel path for wolfram client
KERNEL_PATH = 'D:\\Wolfram Research\\Wolfram Engine\\12.1\WolframKernel.exe'

# parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument('step', help='the step')
parser.add_argument('-g', '--gamma', type=float, help='reciprocal of the gamma value for the set')
parser.add_argument('-d', action='store_true', help='whether to display or save file')
args = parser.parse_args()

file = '../output/values.xlsx'
wb = load_workbook(file)

ws = wb.create_sheet('gamma_g{}s{}'.format(args.gamma, args.step))

# select which file to read from
if args.gamma:
	infile = 'k_gamma_gm.txt'
else:
	infile = 'k_gamma_m.txt'

# access the specified kernel
with WolframLanguageSession(KERNEL_PATH) as session:
	with open(infile, 'r') as f:
		print('Succesfully accessed file ', infile, ', beginning execution...', sep='')
		session.evaluate(wlexpr('s = {};'.format(args.step)))
		if args.gamma:
			session.evaluate(wlexpr('gamma = 1 / {};'.format(args.gamma)))
		session.evaluate(wlexpr(f.read()))
		next = session.evaluate(wlexpr('next'))
		lebesgue = session.evaluate(wlexpr('lebesgue'))
		
	# calculate amount of rows
	rows = 2 ** (int(args.step) + 1)
	
	# create dictionary from keys and values
	results = dict(zip(next, lebesgue))
	
	# write the values
	for i in range(rows):
		ws['A{}'.format(i + 1)].value = next[i]
		ws['B{}'.format(i + 1)].value = lebesgue[i]
	
	# save the file
	wb.save(file)
	
# generate and show plot
plt.plot(next, lebesgue)
plt.title('gamma = {}, step = {}'.format(args.gamma if args.gamma else '-predefined-', args.step))
if args.d:
	plt.show()
else:
	plt.savefig('../output/gamma/gamma_g{}s{}.png'.format(args.gamma if args.gamma else '1 / 3.99', args.step))