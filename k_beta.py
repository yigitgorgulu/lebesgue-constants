import argparse

from wolframclient.evaluation import WolframLanguageSession
from wolframclient.language import wl, wlexpr

from openpyxl import load_workbook, Workbook

import matplotlib.pyplot as plt
import numpy as np

# kernel path for wolfram client
KERNEL_PATH = 'D:\\Wolfram Research\\Wolfram Engine\\12.1\WolframKernel.exe'

# parse command line arguments
parser = argparse.ArgumentParser()
parser.add_argument('a', help='the a value for the set')
parser.add_argument('s', help='the step')
parser.add_argument('-m', action='store_true', help='whether to include m')
parser.add_argument('-d', action='store_true', help='whether to display or save file')
parser.add_argument('-x', help='the xlsx file to access, new one created if not specified')
args = parser.parse_args()

# check whether an xlsx file was provided and open it (or a new one)
if args.x:
	file = args.x
	wb = load_workbook(file)
else:
	file = '../output/values.xlsx'
	wb = load_workbook(file)

# create new sheet for this round
ws = wb.create_sheet('beta_b{}s{}{}'.format(args.a, args.s, 'm' if args.m else ''))

# access the specified kernel
with WolframLanguageSession(KERNEL_PATH) as session:
	# check whether to include xm
	if args.m:
		session.evaluate(wlexpr('a = {};'.format(args.a)))
		
		session.evaluate(wlexpr('cantor[0] = {0, 1};'))
		session.evaluate(wlexpr('cantor[n_] := Join[cantor[n - 1] / a, (a - 1 + cantor[n - 1]) / a];'))
		
		session.evaluate(wlexpr('s = {};'.format(args.s)))
		session.evaluate(wlexpr('set = cantor[s - 1]'))
		set = session.evaluate(wlexpr('N[set]'))
		session.evaluate(wlexpr('next = cantor[s];'))
		next = session.evaluate(wlexpr('N[next]'))
		lebesgue = session.evaluate(wlexpr('Table[N[Sum[Product[If[k == j, 1, Abs[(next[[i]] - set[[k]]) / (set[[j]] - set[[k]])]], {k, 1, 2 ^ s}], {j, 1, 2 ^ s}]], {i, 1, 2 ^ (s + 1)}]'))
		
	else:
		session.evaluate(wlexpr('a = {};'.format(args.a)))
		
		session.evaluate(wlexpr('cantor[0] = {0, 1};'))
		session.evaluate(wlexpr('cantor[n_] := Join[cantor[n - 1] / a, (a - 1 + cantor[n - 1]) / a];'))
		
		session.evaluate(wlexpr('findxm[0]:=1;'))
		session.evaluate(wlexpr('findxm[n_]:=If[Mod[n,2]==1,findxm[n-1]-(a^- n),findxm[n-1]+(a^- n)];'))
		
		session.evaluate(wlexpr('s = {};'.format(args.s)))
		session.evaluate(wlexpr('set = cantor[s - 1]'))
		set = session.evaluate(wlexpr('N[set]'))
		session.evaluate(wlexpr('next = cantor[s];'))
		next = session.evaluate(wlexpr('N[next]'))
		session.evaluate(wlexpr('xm = findxm[s - 1];'))
		session.evaluate(wlexpr('m = Position[set, xm][[1, 1]];'))
		lebesgue = session.evaluate(wlexpr('Table[N[Sum[If[j == m, 0, Product[If[k == m || k == j, 1, Abs[(next[[i]] - set[[k]]) / (set[[j]] - set[[k]])]], {k, 1, 2 ^ s}]], {j, 1, 2 ^ s}]], {i, 1, 2 ^ (s + 1)}]'))
	
	# calculate amount of rows
	rows = 2 ** (int(args.s) + 1)
	
	# create dictionary from keys and values
	results = dict(zip(next, lebesgue))
	
	# write the values
	for i in range(rows):
		ws['A{}'.format(i + 1)].value = next[i]
		ws['B{}'.format(i + 1)].value = lebesgue[i]
	
	# save the file
	wb.save(file)
	
	# generate and show plot
	plt.plot(next, lebesgue)
	plt.title('a = {}, s = {}{}'.format(args.a, args.s, ', xm included' if args.m else ''))
	if args.d:
		plt.show()
	else:
		plt.savefig('../output/beta/beta_a{}s{}{}.png'.format(args.a, args.s, 'm' if args.m else ''))
	